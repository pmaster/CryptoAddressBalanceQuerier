#!/usr/bin/env python3
"""Combine the balances + inbounds CSVs into one Google-Sheets-ready workbook.

    python3 batch/build_workbook.py [--prefix sp4]

Produces batch/output/<prefix>_wallets.xlsx with two tabs:
  Balances  one row per wallet; the inbound aggregates are live SUMIFS/COUNTIFS
            against the Inbounds tab, so they follow any edits or added rows
  Inbounds  one row per qualifying inbound transfer, newest first

Both tabs get a frozen header row and an autofilter for slicing in Sheets.
"""

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "output")

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
TOTAL_FONT = Font(name=FONT, bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
USD = '$#,##0.00;($#,##0.00);-'
AMT = '#,##0.00000000;-#,##0.00000000;-'

# column -> (header label, width, number format)
BAL_COLS = [
    ("wallet", 14, None), ("address", 44, None),
    ("total_usd", 15, USD), ("wallet_tokens_usd", 17, USD), ("defi_usd", 12, USD),
    ("inbound_count", 14, '#,##0'), ("inbound_total_usd", 18, USD),
    ("external_receive_usd", 20, USD),
    ("chain_count", 12, '#,##0'), ("chain_breakdown", 46, None),
    ("top_holdings", 70, None), ("top_defi", 40, None),
    ("history_truncated", 16, None),
]
IN_COLS = [
    ("wallet", 14, None), ("address", 44, None),
    ("datetime_utc", 19, None), ("date_utc", 12, None),
    ("operation_type", 15, None), ("external_receive", 16, None),
    ("internal_transfer", 16, None), ("chain", 14, None),
    ("token", 10, None), ("token_name", 22, None),
    ("amount", 18, AMT), ("price_usd", 12, '#,##0.000000'),
    ("usd_value", 14, USD),
    ("sender", 44, None), ("sender_label", 14, None), ("tx_hash", 68, None),
]


def read_csv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def style_sheet(ws, cols, nrows):
    for idx, (label, width, fmt) in enumerate(cols, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN)
        for r in range(2, nrows + 2):
            c = ws.cell(row=r, column=idx)
            c.font = BODY_FONT
            if fmt:
                c.number_format = fmt
    ws.freeze_panes = "A2"
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{nrows + 1}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prefix", default="sp4")
    args = ap.parse_args()

    bal = read_csv(os.path.join(OUT_DIR, f"{args.prefix}_balances.csv"))
    inb = read_csv(os.path.join(OUT_DIR, f"{args.prefix}_inbounds.csv"))

    wb = Workbook()

    # ---- Inbounds tab (written first so Balances can reference it) ----
    ws_in = wb.active
    ws_in.title = "Inbounds"
    for r, row in enumerate(inb, start=2):
        for idx, (key, _, fmt) in enumerate(IN_COLS, start=1):
            val = row.get(key, "")
            if key in ("amount", "price_usd", "usd_value"):
                val = float(val or 0)
            ws_in.cell(row=r, column=idx, value=val)
    style_sheet(ws_in, IN_COLS, len(inb))

    n_in = len(inb) + 1
    wallet_rng = f"Inbounds!$A$2:$A${n_in}"
    value_rng = f"Inbounds!$M$2:$M${n_in}"
    ext_rng = f"Inbounds!$F$2:$F${n_in}"

    # ---- Balances tab ----
    ws_b = wb.create_sheet("Balances", 0)
    for r, row in enumerate(bal, start=2):
        label = row["wallet"]
        for idx, (key, _, fmt) in enumerate(BAL_COLS, start=1):
            if key == "inbound_count":
                val = f'=COUNTIFS({wallet_rng},$A{r})'
            elif key == "inbound_total_usd":
                val = f'=SUMIFS({value_rng},{wallet_rng},$A{r})'
            elif key == "external_receive_usd":
                val = f'=SUMIFS({value_rng},{wallet_rng},$A{r},{ext_rng},"TRUE")'
            else:
                val = row.get(key, "")
                if key in ("total_usd", "wallet_tokens_usd", "defi_usd"):
                    val = float(val or 0)
                elif key == "chain_count":
                    val = int(val or 0)
            ws_b.cell(row=r, column=idx, value=val)
    style_sheet(ws_b, BAL_COLS, len(bal))

    # totals row
    tr = len(bal) + 2
    ws_b.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
    for idx, (key, _, fmt) in enumerate(BAL_COLS, start=1):
        if key in ("total_usd", "wallet_tokens_usd", "defi_usd", "inbound_count",
                   "inbound_total_usd", "external_receive_usd"):
            letter = get_column_letter(idx)
            c = ws_b.cell(row=tr, column=idx, value=f"=SUM({letter}2:{letter}{tr - 1})")
            c.font = TOTAL_FONT
            c.number_format = fmt
            c.border = Border(top=THIN)

    out = os.path.join(OUT_DIR, f"{args.prefix}_wallets.xlsx")
    wb.save(out)
    print(f"wrote {out}  ({len(bal)} wallets, {len(inb)} inbound transfers)")


if __name__ == "__main__":
    main()
